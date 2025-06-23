import openpyxl
from openpyxl import Workbook
from datetime import datetime

class ExcelManager:
    def __init__(self, filename='data_collector.xlsx'):
        self.filename = filename
        try:
            self.wb = openpyxl.load_workbook(filename=self.filename)
        except FileNotFoundError:
            self.wb = Workbook()
            self.setup_sheets()
            self.save()

        self.ws_users = self.wb['users']
        self.ws_posts = self.wb['posts']
        self.ws_friends = self.wb['friends']
        self.ws_friend_requests = self.wb['friend request updates']
        self.ws_comments = self.wb['comments']
        self.ws_messages = self.wb['messages']

    def setup_sheets(self):
        ws = self.wb.active
        ws.title = 'users'
        self.wb.create_sheet('posts')
        self.wb.create_sheet('friends')
        self.wb.create_sheet('friend request updates')
        self.wb.create_sheet('comments')
        self.wb.create_sheet('messages')

    def save(self):
        self.wb.save(self.filename)

    def get_last_row(self, sheet):
        return sheet.max_row

    def get_last_column(self, sheet):
        return sheet.max_column


class UserManager:
    def __init__(self, excel_manager):
        self.excel_manager = excel_manager
        self.users_sheet = self.excel_manager.ws_users
        self.logged_in_user = None

    def sign_up(self, name, password, user_id, email, address, nickname):
        last_row = self.excel_manager.get_last_row(self.users_sheet)
        for row in range(2, last_row + 1):
            if self.users_sheet.cell(row=row, column=5).value == name:
                print("User already exists.")
                return False

        # Add new user to the Excel sheet
        self.users_sheet.cell(row=last_row + 1, column=5, value=name)
        self.users_sheet.cell(row=last_row + 1, column=3, value=password)
        self.users_sheet.cell(row=last_row + 1, column=4, value=user_id)
        self.users_sheet.cell(row=last_row + 1, column=2, value=email)
        self.excel_manager.save()
        print(f"User {name} signed up successfully!")
        return True

    def login(self, name, password):
        last_row = self.excel_manager.get_last_row(self.users_sheet)
        for row in range(2, last_row + 1):
            if (self.users_sheet.cell(row=row, column=5).value == name and 
                self.users_sheet.cell(row=row, column=3).value == password):
                print(f"User {name} logged in successfully!")
                self.logged_in_user = name
                return True
        print("Invalid login credentials.")
        return False

    def logout(self):
        if self.logged_in_user:
            print(f"User {self.logged_in_user} logged out.")
            self.logged_in_user = None
        else:
            print("No user is currently logged in.")


class FriendManager:
    def __init__(self, excel_manager, user_manager):
        self.excel_manager = excel_manager
        self.user_manager = user_manager
        self.friends_sheet = self.excel_manager.ws_friends
        self.friend_requests_sheet = self.excel_manager.ws_friend_requests

    def send_friend_request(self, friend_name):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.friend_requests_sheet)
        self.friend_requests_sheet.cell(row=last_row + 1, column=1, value=self.user_manager.logged_in_user)
        self.friend_requests_sheet.cell(row=last_row + 1, column=3, value=friend_name)
        self.friend_requests_sheet.cell(row=last_row + 1, column=2, value='friend request')
        self.excel_manager.save()
        print(f"Friend request sent to {friend_name}.")
        return True

    def view_friend_requests(self):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.friend_requests_sheet)
        print(f"\n--- Friend Requests for {self.user_manager.logged_in_user} ---")
        for row in range(2, last_row + 1):
            if self.friend_requests_sheet.cell(row=row, column=3).value == self.user_manager.logged_in_user:
                sender = self.friend_requests_sheet.cell(row=row, column=1).value
                request_type = self.friend_requests_sheet.cell(row=row, column=2).value
                print(f"From {sender}: {request_type} (ID: {row})")

    def accept_friend_request(self, request_id):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        row = int(request_id)
        if row <= 1 or row > self.excel_manager.get_last_row(self.friend_requests_sheet):
            print("Invalid request ID.")
            return False

        sender = self.friend_requests_sheet.cell(row=row, column=1).value
        recipient = self.friend_requests_sheet.cell(row=row, column=3).value
        request_type = self.friend_requests_sheet.cell(row=row, column=2).value

        if recipient != self.user_manager.logged_in_user or request_type != 'friend request':
            print("This request does not belong to you or is not a friend request.")
            return False

        # Add to friends list
        last_row = self.excel_manager.get_last_row(self.friends_sheet)
        self.friends_sheet.cell(row=last_row + 1, column=1, value=sender)
        self.friends_sheet.cell(row=last_row + 1, column=2, value=self.user_manager.logged_in_user)

        self.friends_sheet.cell(row=last_row + 2, column=1, value=self.user_manager.logged_in_user)
        self.friends_sheet.cell(row=last_row + 2, column=2, value=sender)

        # Remove from friend requests
        self.friend_requests_sheet.delete_rows(row)
        self.excel_manager.save()
        print(f"Friend request from {sender} accepted.")
        return True


class PostManager:
    def __init__(self, excel_manager, user_manager):
        self.excel_manager = excel_manager
        self.user_manager = user_manager
        self.posts_sheet = self.excel_manager.ws_posts

    def create_post(self, content):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.posts_sheet)
        self.posts_sheet.cell(row=last_row + 1, column=1, value=self.user_manager.logged_in_user)
        self.posts_sheet.cell(row=last_row + 1, column=2, value=content)
        self.excel_manager.save()
        print(f"Post created by {self.user_manager.logged_in_user}.")
        return True

    def view_posts(self):
        last_row = self.excel_manager.get_last_row(self.posts_sheet)
        print("\n--- Posts ---")
        for row in range(2, last_row + 1):
            user = self.posts_sheet.cell(row=row, column=1).value
            content = self.posts_sheet.cell(row=row, column=2).value
            print(f"{user}: {content}")


class CommentManager:
    def __init__(self, excel_manager, user_manager):
        self.excel_manager = excel_manager
        self.user_manager = user_manager
        self.comments_sheet = self.excel_manager.ws_comments

    def add_comment(self, post_id, comment):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.comments_sheet)
        self.comments_sheet.cell(row=last_row + 1, column=1, value=post_id)
        self.comments_sheet.cell(row=last_row + 1, column=2, value=self.user_manager.logged_in_user)
        self.comments_sheet.cell(row=last_row + 1, column=3, value=comment)
        self.excel_manager.save()
        print(f"Comment added to post {post_id}.")
        return True

    def view_comments(self, post_id):
        last_row = self.excel_manager.get_last_row(self.comments_sheet)
        print(f"\n--- Comments on Post {post_id} ---")
        for row in range(2, last_row + 1):
            if self.comments_sheet.cell(row=row, column=1).value == post_id:
                user = self.comments_sheet.cell(row=row, column=2).value
                comment = self.comments_sheet.cell(row=row, column=3).value
                print(f"{user}: {comment}")


class MessageManager:
    def __init__(self, excel_manager, user_manager):
        self.excel_manager = excel_manager
        self.user_manager = user_manager
        self.messages_sheet = self.excel_manager.ws_messages

    def send_message(self, recipient, message):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.messages_sheet)
        self.messages_sheet.cell(row=last_row + 1, column=1, value=self.user_manager.logged_in_user)
        self.messages_sheet.cell(row=last_row + 1, column=2, value=recipient)
        self.messages_sheet.cell(row=last_row + 1, column=3, value=message)
        self.excel_manager.save()
        print(f"Message sent to {recipient}.")
        return True

    def view_inbox(self):
        if not self.user_manager.logged_in_user:
            print("You need to log in first.")
            return False

        last_row = self.excel_manager.get_last_row(self.messages_sheet)
        print(f"\n--- Inbox for {self.user_manager.logged_in_user} ---")
        for row in range(2, last_row + 1):
            sender = self.messages_sheet.cell(row=row, column=1).value
            recipient = self.messages_sheet.cell(row=row, column=2).value
            message = self.messages_sheet.cell(row=row, column=3).value
            if recipient == self.user_manager.logged_in_user:
                print(f"From {sender}: {message}")
        return True


def main():
    excel_manager = ExcelManager()
    user_manager = UserManager(excel_manager)
    friend_manager = FriendManager(excel_manager, user_manager)
    post_manager = PostManager(excel_manager, user_manager)
    comment_manager = CommentManager(excel_manager, user_manager)
    message_manager = MessageManager(excel_manager, user_manager)

    while True:
        print("\n1. Sign Up\n2. Log In\n3. Exit")
        choice = input("Enter your choice: ")

        if choice == "1":
            name = input("Enter your name: ")
            password = input("Enter your password: ")
            user_id = input("Enter your ID: ")
            email = input("Enter your email: ")
            address = input("Enter your address: ")
            nickname = input("Enter your nickname: ")
            user_manager.sign_up(name, password, user_id, email, address, nickname)

        elif choice == "2":
            name = input("Enter your name: ")
            password = input("Enter your password: ")
            if user_manager.login(name, password):
                while True:
                    print("\n1. Send Friend Request\n2. View Friend Requests\n3. Accept Friend Request\n4. Create Post\n5. View Posts\n6. Add Comment\n7. View Comments\n8. Send Private Message\n9. View Inbox\n10. Log Out")
                    user_choice = input("Enter your choice: ")

                    if user_choice == "1":
                        friend_name = input("Enter the friend's name: ")
                        friend_manager.send_friend_request(friend_name)

                    elif user_choice == "2":
                        friend_manager.view_friend_requests()

                    elif user_choice == "3":
                        request_id = input("Enter the ID of the request to accept: ")
                        friend_manager.accept_friend_request(request_id)

                    elif user_choice == "4":
                        content = input("Enter post content: ")
                        post_manager.create_post(content)

                    elif user_choice == "5":
                        post_manager.view_posts()

                    elif user_choice == "6":
                        post_id = input("Enter post ID to comment on: ")
                        comment = input("Enter your comment: ")
                        comment_manager.add_comment(post_id, comment)

                    elif user_choice == "7":
                        post_id = input("Enter post ID to view comments: ")
                        comment_manager.view_comments(post_id)

                    elif user_choice == "8":
                        recipient = input("Enter the recipient's name: ")
                        message = input("Enter your message: ")
                        message_manager.send_message(recipient, message)

                    elif user_choice == "9":
                        message_manager.view_inbox()

                    elif user_choice == "10":
                        user_manager.logout()
                        break

                    else:
                        print("Invalid choice, please try again.")

        elif choice == "3":
            break

        else:
            print("Invalid choice, please try again.")


if __name__ == "__main__":
    main()
