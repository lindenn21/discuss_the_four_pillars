import openpyxl
from openpyxl import Workbook
def sheet_creation():
    wb = Workbook()
    ws = wb.create_sheet('sheet1')
    ws.title = 'users'
    ws = wb.create_sheet('sheet2')
    ws.title = 'posts'
    ws = wb.create_sheet('sheet3')
    ws.title = 'friends'
    ws=wb.create_sheet('sheet4')
    ws.title='friend request updates'
    wb.save(filename='data_collector.xlsx')
    ws1 = wb['users']
    ws2 = wb['posts']
    ws3 = wb['friends']
    ws4=wb['friend request updates']
    ws1.cell(1, 7, value='facebook current record')
    ws1.cell(2, 1, value='None')
    ws1.cell(2, 2, value='email')
    ws1.cell(2, 3, value='PASSWORD')
    ws1.cell(2, 4, value='id')
    ws1.cell(2, 5, value='name')
    ws3.cell(1,8,value='PEOPLE WITH FRIENDS')
    ws3.cell(2,1,value='NAME')
    ws3.cell(2, 3, value='friends name of people')
    ws4.cell(1,8,value='updates for you as a reminder')
    ws4.cell(2, 1, value='NAMES')
    ws4.cell(2, 2, value='notification')
    ws4.cell(2, 3, value='request list')
    ws2.cell(1,5,value='what is in your mind')
    ws2.cell(1,1,value='name')
    ws2.cell(1,2,value='text post')
    ws2.cell(1,3,value='integr post')
    ws2.cell(1,4,value='COMMENTS')
    ws2.cell(1,5,value='REACT POST')



    lastrow = ws1.max_row
    lastrow3=ws3.max_row
    lastrow4=ws4.max_row
    lastrow2=ws2.max_row
    lastcolumn3 = ws3.max_column
    wb.save('data_collector.xlsx')
wb=openpyxl.load_workbook(filename='data_collector.xlsx')
ws1 = wb['users']
ws2 = wb['posts']
ws3 = wb['friends']
ws4=wb['friend request updates']
lastrow = ws1.max_row
lastrow3=ws3.max_row
lastrow4=ws4.max_row
lastrow2=ws2.max_row
lastcolumn3=ws3.max_column

class data:
    paswordlist = []
    namelist = []
    addingreference = []
    sender=0
    loginlist = []
    list5 = []
    savepreveious=0
    k = 0
    p = 0
    requestlist = []

    store = 0

    def __init__(self, name, pasword,id,email,address,nick_name):
        self.name = name
        self.nick_name=nick_name
        self.address=address
        self.id=id
        self.email=email
        self.pasword = pasword
        self.friendlist = []
        self.sign = False
        self.LOGIN = False
        self.LOGOUT = False
        data.signup(self)
    def signup(self):
        # print(self.name,self.pasword)
        data.paswordlist.append(self.pasword)
        data.namelist.append(self.name)
        flag = True
        a=2
        while ws1.cell(row=a,column=5).value != None:
            if ws1.cell(row=a,column=5).value == self.name:
                flag = False
                self.sign=True
                break
            a+=1
        if flag == True:
            ws1.cell(row=lastrow+1,column=5,value=self.name)
            ws1.cell(row=lastrow+1,column=3,value=self.pasword)
            ws1.cell(row=lastrow+1,column=4,value=self.id)
            ws1.cell(row=lastrow+1,column=2,value=self.email)
            ws3.cell(row=lastrow + 1, column=1, value=self.name)

            wb.save('data_collector.xlsx')
            self.sign = True
        else:
            pass
    def storereference(data):
        data.addingreference.append(data)
        # print(data.addingreference)
    def loginn(self):
        if self.sign == True:
            nroll = str(input("give name to log in"))
            verifypassword = int(input("give password to log in"))
            if nroll==self.name and verifypassword==self.pasword:
                print("you are logged in successfully")
                self.LOGIN = True
                data.loginlist.append(self.name)
                ws1.cell(row=lastrow+1,column=7,value=self.name)
                return self
            else:
                print('invalid input ')
    def editbiodata(self):
        if self.LOGIN==True:
            CHANGE=int(input('press one to change email,press 2 to change nickname,press 3 for address,'))
            if CHANGE==1:
                self.email=input('give new email you want ')
            if CHANGE==2:
                self.nick_name=input('give nick name')
            if CHANGE==3:
                self.address=input('new address')
        else:
            print('plz login first')
    def searching(self):
        if self.LOGIN==True:
            search=input('give name of person you want to search')
            for i in data.addingreference:
                if search==i.name:
                    print(i.name, 'is present at ' ,data.addingreference.index(i))
    def request(self):
        if self.LOGIN == True:
            # print('if you want to send friend request,press 1')
            # print('if you want to check and reply for friend request,press 2')
            # decision=int(input('give number '))
            # if decision==1:
            data.p = input("give the name of people you want to add")
            for i in data.addingreference:
                if data.p == i.name and self.name in data.namelist:
                    data.store = i
                    print(data.store)
                    data.sender=self.name
                    acceptor=data.addingreference.index(i)
                    requester=data.namelist.index(self.name)
                    print("request snd")
                    data.savepreveious=lastrow4
                    ws4.cell(row=1+lastrow4, column=1, value=self.name)
                    ws4.cell(row=1+lastrow4, column=3, value=i.name)
                    ws4.cell(row=1+lastrow4, column=2, value='friend request')
                    wb.save('data_collector.xlsx')
                    # s = add.decide()
                    print(requester)
                    print(acceptor)
                    data.requestlist.append(data.p)
    def react(self):
        acceptor=None
        requester=None
        sender_refrence=0
        for i in range(1,5000):

            if self.name==ws4.cell(row=i,column=3).value:
                acceptor=self.name
                print(acceptor,'wanted us')
                sender_refrence=i+1
                print(sender_refrence)
                requester=ws4.cell(row=i,column=1).value
                print(requester,'wanted us ')
        # print(data.store)
        # acceptor = data.addingreference.index(data.store)
        print(sender_refrence,'is here')
        # requester = data.namelist.index(data.sender)
        if self.LOGIN==True:
            s=add.decide(self)
            if s == 1:
                # i.friendlist.append(self.name)
                # print(data.store.name,data.sender)
                # ws4.cell(row=data.savepreveious+1, column=1, value=None)
                # ws4.cell(row=data.savepreveious+1, column=3, value=None)
                # ws4.cell(row=data.savepreveious+1, column=2, value=None)
                ws3.cell(row=sender_refrence+1, column=lastcolumn3+1, value=requester)
                ws3.cell(row=sender_refrence+1, column=lastcolumn3+2, value=self.name)
                # ws3.cell(row=acceptor + 3, column=lastrow3 + 1, value=i.name)
                # ws3.cell(row=requester + 3, column=2, value=i.name)
                # self.friendlist.append(self.name)
                wb.save('data_collector.xlsx')
                print('HY')
    def messege(self):
        if self.LOGIN==True :
            u=int(input('press one to send mssage and 2 to check'))
            if u==1:
                print('text your message .....')
                for i in range (1,30):
                    if self.name==ws3.cell(row=i,column=1).value:
                        indexs=i
                print(indexs)
                msg=input('type.....')
                to=input('give name to whom you message.....')
                sha=[]
                a=2
                while ws3.cell(indexs,a).value!=None:
                    sha.append(ws3.cell(indexs,a).value)
                    a+=1
                print(sha)
                for i in sha:
                    if to==i:
                        print('sent message')
                        ws4.cell(row=lastrow4+1, column=7, value=self.name)
                        ws4.cell(row=lastrow4+1, column=5, value=to)
                        ws4.cell(row=lastrow4+1, column=2, value='chatting')
                        ws4.cell(row=lastrow4 + 1, column=6, value=msg)
                        wb.save('data_collector.xlsx')
                        add.receivemsg(self)
                    if to not in sha:
                        print('user does not exsist...')
            if u==2:
                for i in range(1, 2000):
                    if self.name == ws4.cell(row=i, column=3).value:
                        add.receivemsg(self)

        else:
            print('login first plzz')
    def loggedout(self):
        if self.LOGIN==True:
            self.LOGOUT=True
            self.LOGIN=False
            for i in range(1,20):
                if self.name== ws1.cell(row=i, column=7).value:
                    ws1.cell(row=i, column=7,value=None)
                    print('you are logged out')

class add:
    # def __init__(self):
    # self.friendlist=[]

    def decide(self):
        d=0
        print(ws4.cell(row=3,column=3).value,'hhh')
        flag = False
        for i in range(1,2000):
            if self.name == ws4.cell(row=i,column=3).value:
                print(self.name)
                d=i
                flag = True
        if flag == True:

            print('notification')
            u = int(input("give your choice about ffriend request you have received "))
            # addingreference.append(data)
            if u == 1:
                print('now you are friends on facebook')
                # data.store..append(data.store.name)
                ws4.cell(row=d, column=3).value=None
                ws4.cell(row=d, column=1).value=None
                ws4.cell(row=d, column=2).value=None
                return 1
            else:
                return 2
    def receivemsg(self):
        flag = False
        for i in range(1, 2000):
            if self.name == ws4.cell(row=i, column=3).value:
                print(self.name)
                d = i
                flag = True
        if flag == True:

            print("you have a message ")
            check=int(input('if you want to reply press 1'))
            if check==1:
                print('message seen')
                data.messege(self)
                return 'message seen'
            if check==2:
                print('message seen but didnot want to reply')
class sharepost:
    # data.loginn(self)


    def post(self):
        data.loginn(self)
        print('what is in your mind ')
        post=input('type ....')
        print(self.name)
        place=data.namelist.index(self.name)
        ws2.cell(row=lastrow2+1,column=1,value=self.name)
        ws2.cell(row=lastrow2 + 1, column=2, value=post)
        print('if you want to hide post,press one ')
        option=int(input('press one to use option'))
        if option==1:

            hide=input('give name of people you want to hide post from ')
            ws3.cell(row=lastrow3+1,column=8)
            wb.save('data_collector.xlsx')
        else:

            for i in range(1,6):
                # while ws3.cell(row=place+1,column=i).value!=None:
                k=ws3.cell(row=place+3,column=i).value
                if k == ws3.cell(row=i,column=8):

                    print(k)
                else:
                    ws4.cell(row=i + 1, column=1, value=self.name)
                    ws4.cell(row=i + 1, column=2, value='post by friend')
                    ws4.cell(row=i + 1, column=3, value=k)
                    wb.save('data_collector.xlsx')
class comments:
    def takecoment(self):
        data.loginn(self)
        print(self.name)
        print(ws4.cell(row=3, column=3).value)
        for i in range(2, 10):
            if ws4.cell(row=i, column=3).value == self.name:
                print('to which post you want to comment')
                choice=int(input('press the number to which you want to comment'))+1
                comentpas=input('give your reaction')
                ws2.cell(row=choice, column=4, value=comentpas)
                ws2.cell(row=choice, column=5, value=self.name)
                wb.save('data_collector.xlsx')
                print(ws2.cell(row=choice, column=2).value)
                print(comentpas)
    def searchpostbyword(self):
        if data.loginlist!=None:
            search=input('some key words of post')
            for i in range(2,3):
                postsearch=(ws2.cell(row=i, column=2)).value
                # print(postsearch)
                postsearch=postsearch.split(" ")
                print(postsearch)
                if search==postsearch[i]:
                    print((ws2.cell(row=i, column=2)).value)
                    break
                # print('no notification ')
# print(sheet_creation())
second = data('umer', 2,3344,'umer123','karachi','umer')
# # second.signup()
first = data('kashif', 8,1122,'k516','mianwali','kashi')
third=data('sarib',420,5566,'gardezi125','kashmir','maalik')
# first.signup()
# # second.signup()
# third.loginn()
second.loginn()
# second.loginn()
data.storereference(first)
data.storereference(second)
data.storereference(third)
# second.request()
# ]
# first.react()
# second.messege()
# sharepost.post(second)
# comments.takecoment(third)
#list(kashi.split(" "))
comments.searchpostbyword(second)